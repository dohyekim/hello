import openpyxl

book = openpyxl.workbook()
sheet1 = book.active
sheet1.title = "첫번째 Sheet"
sheet1.cell(row = 1, column = 1).value = "Company name"

sheet2 = book.create_sheet()
sheet2.title = "두번째 Sheet"
sheet2.cell(row = 1, column =1).value = "NO url"

book.save("./output.xlsx")